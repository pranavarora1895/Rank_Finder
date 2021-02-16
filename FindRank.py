""" Merit List and School Rank Calculator - By Pranav Arora

    Please follow some simple instructions:
        1. In your excel file write the roll numbers (or names) starting from cell A2
        2. Write the marks starting from cell B2
        3. Use only column A and B. The rank will appear in column C.
"""
import pythoncom
from openpyxl import load_workbook
import win32com.client
from tkinter import Tk
from tkinter.filedialog import askopenfilename
from tkinter import messagebox
# ------------------------------------------------------------------------------------------------------
from openpyxl.utils.exceptions import InvalidFileException

print('''Merit List and School Rank Calculator - By Pranav Arora

    Please follow some simple instructions:
        1. In your excel file write the roll numbers (or names) starting from cell A2
        2. Write the marks starting from cell B2
        3. Use only column A and B. The rank will appear in column C.
        4. Make sure to close the excel file before running the tool. 
        ''')
# -------------------------------------------------------------------------------------------------------
try:

    def get_excel_path():
        """Getting the file (GUI)"""
        Tk().withdraw()
        return askopenfilename(filetypes=[('Excel Files', '*.csv;*.xlsx;*.xlsm;*.xltx;*.xltm')],
                               title='Select your Excel File and Click Open')


    filename = get_excel_path()
    if '\\' in filename:
        filename = filename.replace('\\', '/')
    if '.xlsx' not in filename:
        filename += '.xlsx'
    generic = ""
    last_slash = filename.rfind('/')
    for term in range(last_slash):
        generic += filename[term]
# -------------------------------------------------------------------------------------------------------
    # loading the workbook
    wb2_pyxl = load_workbook(filename)
    ws_pyxl = wb2_pyxl.active
    workbook_elements_list = [0]

    no_of_students = -1
    for cell in ws_pyxl['A']:
        no_of_students += 1

    for row in ws_pyxl.iter_rows(min_row=2, max_col=2, max_row=no_of_students+1, values_only=True):
        workbook_elements_list.append(row)
    #    print(row)
    #print(workbook_elements_list)
    roll_marks_map={}
    for i in range(1, len(workbook_elements_list)):
        roll_marks_map[ws_pyxl.cell(row=i+1, column=1).value] = workbook_elements_list[i][1]
    #print(roll_marks_map)

    # reversing the key_value pair
    roll_marks_map = {value:key for key, value in roll_marks_map.items()}
    #print(roll_marks_map)
# --------------------------------------------------------------------------------------------------------
    # sorting marks
    excel = win32com.client.Dispatch("Excel.Application")

    wb_win = excel.Workbooks.Open(filename)
    ws_win = wb_win.Worksheets('Sheet1')

    ran = 'B2:B'+str(no_of_students+1)
    ws_win.Range(ran).Sort(Key1=ws_win.Range('B1'), Order1=2, Orientation=1)
    sorted_file = generic + '/py_sorted_file.xlsx'
    sorted_file = sorted_file.replace('/', '\\')
    wb_win.SaveAs(sorted_file)
    excel.Application.Quit()
# ----------------------------------------------------------------------------------------------------------
    sorted_file = sorted_file.replace('\\','/')

    # loading roll no from the roll_marks_map for the specific marks
    wb3_pyxl = load_workbook(sorted_file)
    ws2_pyxl = wb3_pyxl.active
    ws2_pyxl['C1'] = 'Rank'
    school_rank = {}
    # Creating Merit List
    for row in range(2, len(workbook_elements_list) + 1):
            marks = ws2_pyxl.cell(row=row,column=2).value
            ws2_pyxl.cell(row=row,column=1).value = roll_marks_map.get(marks)
            ws2_pyxl.cell(row=row,column=3).value = row-1
            school_rank[marks] = row-1
    #print(school_rank)
    merit_list_location = generic + '/Merit_List.xlsx'
    wb3_pyxl.save(merit_list_location)

    messagebox.showinfo("Merit List Created", f'Merit List Created at {merit_list_location}')
# ----------------------------------------------------------------------------------------------------------
    # Creating School Rank
    wb4_pyxl = load_workbook(filename)
    ws3_pyxl = wb4_pyxl.active
    ws3_pyxl['C1'] = 'Rank'
    for rank in range(2,len(workbook_elements_list)+1):
        marks_2 = ws3_pyxl.cell(row=rank,column=2).value
    #    print(marks_2)
        ws3_pyxl.cell(row=rank,column=3).value = school_rank.get(marks_2)
    #    print(ws3_pyxl.cell(row=rank,column=3).value)

    wb4_pyxl.save(filename)

    messagebox.showinfo("Rank Created", f'School Rank Created in your original file: {filename} \n Please ignore '
                            f'py_sorted_file.xlsx')

except FileNotFoundError as e:
    print(f'Error: {e.errno}\nFile Not Found')
except OSError as o:
    print(f'Error: {o.errno}\nInvalid Entry')
except ValueError as v:
    print('ValueError')
except pythoncom.com_error as error:
    print('Invalid Cell Number')
    print(error)
    print(vars(error))
    print(error.args)
    hr, msg, exc, arg = error.args
except AttributeError as a:
    print('Oops!!! Close your excel file and run the tool again')
except InvalidFileException:
    print('Operation Cancelled')


# ------------------x-------------------------------------x--------------------------------x-----------------

